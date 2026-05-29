import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.catalogo.models import Asignatura, Carrera, TipoEquipo
from apps.inventario.models import Unidad

CODIGO_ACTIVO_RE = re.compile(r"\b\d{2}-\d+\b")
CODIGO_SIN_GUION_RE = re.compile(r"\b\d{10,}\b")
CODIGO_ASIGNATURA_RE = re.compile(r"[A-Z]{4}\d{2}")

HEADER_ALIASES = {
    "nombre": (
        "nombre equipo",
        "nombre del equipo",
        "equipo",
        "descripcion equipo",
        "descripcion",
        "nombre",
    ),
    "codigo": (
        "codigo inventario",
        "codigo de inventario",
        "codigo activo",
        "cod activo",
        "inventario",
    ),
    "cantidad": ("cantidad total", "cantidad", "total"),
    "asignaturas": ("asignatura", "asignaturas", "codigo asignatura"),
    "carreras": ("carrera", "carreras"),
    "especificacion": ("especificacion", "especificaciones", "detalle"),
}

FALLBACK_COLUMNS = {
    "nombre": 0,
    "cantidad": 1,
    "codigo": 2,
    "asignaturas": 3,
    "carreras": 4,
    "especificacion": 5,
}


@dataclass
class ImportSummary:
    tipos_creados: int = 0
    tipos_actualizados: int = 0
    unidades_con_codigo: int = 0
    unidades_sin_codigo: int = 0
    filas_revision: int = 0
    warnings: list[str] = field(default_factory=list)
    _tipos_actualizados_ids: set[int] = field(default_factory=set)

    def mark_tipo_actualizado(self, tipo_equipo: TipoEquipo) -> None:
        if tipo_equipo.pk not in self._tipos_actualizados_ids:
            self.tipos_actualizados += 1
            self._tipos_actualizados_ids.add(tipo_equipo.pk)


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_positive_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, int):
        return max(value, 0)
    if isinstance(value, float):
        return max(int(value), 0)
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else 0


def normalize_inventory_codes(value: Any) -> tuple[list[str], bool]:
    text = clean_text(value)
    if not text or text == "???":
        return [], bool(text == "???" or not text)

    codes = CODIGO_ACTIVO_RE.findall(text)
    remaining_text = CODIGO_ACTIVO_RE.sub("", text)
    requires_review = bool(codes and re.sub(r"\s+", "", remaining_text))

    for raw_code in CODIGO_SIN_GUION_RE.findall(text):
        normalized_code = f"{raw_code[:2]}-{raw_code[2:]}"
        if normalized_code not in codes:
            codes.append(normalized_code)
        requires_review = True

    if not codes:
        requires_review = True

    return codes, requires_review


def parse_subject_codes(*values: Any) -> set[str]:
    text = " ".join(clean_text(value).upper() for value in values if clean_text(value))
    return set(CODIGO_ASIGNATURA_RE.findall(text))


def build_column_map(rows: list[tuple[Any, ...]]) -> dict[str, int]:
    normalized_aliases = {
        field: {normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    column_map: dict[str, int] = {}

    for row in rows[:3]:
        for index, value in enumerate(row):
            header = normalize_header(value)
            if not header:
                continue
            for field_name, aliases in normalized_aliases.items():
                if field_name not in column_map and header in aliases:
                    column_map[field_name] = index

    return {**FALLBACK_COLUMNS, **column_map}


def get_cell(row: tuple[Any, ...], column_map: dict[str, int], field_name: str) -> Any:
    index = column_map[field_name]
    return row[index] if index < len(row) else None


class Command(BaseCommand):
    help = "Importa inventario desde una planilla .xlsx hacia TipoEquipo y Unidad."

    def add_arguments(self, parser):
        parser.add_argument(
            "ruta", type=str, help="Ruta al archivo .xlsx de inventario"
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Ejecuta la importación y revierte los cambios al finalizar",
        )

    def handle(self, *args, **options):
        path = Path(options["ruta"])
        dry_run = options["dry_run"]

        if not path.exists():
            raise CommandError(f"No existe el archivo: {path}")
        if path.suffix.lower() != ".xlsx":
            raise CommandError("El archivo debe tener extensión .xlsx")

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        column_map = build_column_map(rows)

        with transaction.atomic():
            summary = self.import_rows(rows[3:], column_map)
            if dry_run:
                transaction.set_rollback(True)

        for warning in summary.warnings:
            self.stdout.write(self.style.WARNING(warning))

        dry_run_prefix = "[dry-run] " if dry_run else ""
        self.stdout.write(
            self.style.SUCCESS(
                f"{dry_run_prefix}tipos creados: {summary.tipos_creados}; "
                f"tipos actualizados: {summary.tipos_actualizados}; "
                f"unidades con código: {summary.unidades_con_codigo}; "
                f"unidades sin código: {summary.unidades_sin_codigo}; "
                f"filas marcadas a revisar: {summary.filas_revision}"
            )
        )

    def import_rows(
        self,
        rows: list[tuple[Any, ...]],
        column_map: dict[str, int],
    ) -> ImportSummary:
        summary = ImportSummary()
        seen_codes: set[str] = set()

        for row_number, row in enumerate(rows, start=4):
            nombre = clean_text(get_cell(row, column_map, "nombre"))
            if not nombre:
                continue

            codigo_cell = get_cell(row, column_map, "codigo")
            cantidad_total = parse_positive_int(get_cell(row, column_map, "cantidad"))
            especificacion = clean_text(get_cell(row, column_map, "especificacion"))
            codigos, row_requires_review = normalize_inventory_codes(codigo_cell)
            cantidad_sin_codigo = max(cantidad_total - len(codigos), 0)
            if row_requires_review and not codigos and cantidad_sin_codigo == 0:
                cantidad_sin_codigo = 1
            row_requires_review = row_requires_review or cantidad_sin_codigo > 0

            tipo_equipo, tipo_created = TipoEquipo.objects.get_or_create(nombre=nombre)
            if tipo_created:
                summary.tipos_creados += 1

            tipo_changed = self.update_tipo_equipo(
                tipo_equipo,
                cantidad_total,
                especificacion,
            )
            subject_codes = parse_subject_codes(
                get_cell(row, column_map, "asignaturas"),
                get_cell(row, column_map, "carreras"),
            )
            relations_changed = self.link_subjects(tipo_equipo, subject_codes)
            if not tipo_created and (tipo_changed or relations_changed):
                summary.mark_tipo_actualizado(tipo_equipo)

            row_marked = False
            for codigo in codigos:
                if (
                    codigo in seen_codes
                    or Unidad.objects.filter(
                        codigo_activo=codigo,
                    ).exists()
                ):
                    summary.warnings.append(
                        f"Fila {row_number}: código duplicado omitido ({codigo})."
                    )
                    row_marked = True
                    continue

                Unidad.objects.create(
                    tipo_equipo=tipo_equipo,
                    codigo_activo=codigo,
                    requiere_revision=row_requires_review,
                )
                seen_codes.add(codigo)
                summary.unidades_con_codigo += 1

            created_without_code = self.ensure_units_without_code(
                tipo_equipo,
                cantidad_sin_codigo,
            )
            summary.unidades_sin_codigo += created_without_code
            row_marked = row_marked or row_requires_review
            if row_marked:
                summary.filas_revision += 1

        return summary

    def update_tipo_equipo(
        self,
        tipo_equipo: TipoEquipo,
        cantidad_total: int,
        especificacion: str,
    ) -> bool:
        changed = False
        if cantidad_total and tipo_equipo.cantidad_necesaria != cantidad_total:
            tipo_equipo.cantidad_necesaria = cantidad_total
            changed = True
        if especificacion and tipo_equipo.especificacion != especificacion:
            tipo_equipo.especificacion = especificacion
            changed = True
        if changed:
            tipo_equipo.save(update_fields=["cantidad_necesaria", "especificacion"])
        return changed

    def link_subjects(self, tipo_equipo: TipoEquipo, subject_codes: set[str]) -> bool:
        changed = False
        for subject_code in subject_codes:
            asignatura, _ = Asignatura.objects.get_or_create(
                codigo=subject_code,
                defaults={"nombre": subject_code},
            )
            carrera, _ = Carrera.objects.get_or_create(nombre=subject_code[:4])
            if not tipo_equipo.asignaturas.filter(pk=asignatura.pk).exists():
                tipo_equipo.asignaturas.add(asignatura)
                changed = True
            if not tipo_equipo.carreras.filter(pk=carrera.pk).exists():
                tipo_equipo.carreras.add(carrera)
                changed = True
        return changed

    def ensure_units_without_code(
        self,
        tipo_equipo: TipoEquipo,
        requested_count: int,
    ) -> int:
        if requested_count <= 0:
            return 0

        existing_count = Unidad.objects.filter(
            tipo_equipo=tipo_equipo,
            codigo_activo__isnull=True,
            requiere_revision=True,
        ).count()
        units_to_create = max(requested_count - existing_count, 0)
        Unidad.objects.bulk_create(
            Unidad(tipo_equipo=tipo_equipo, requiere_revision=True)
            for _ in range(units_to_create)
        )
        return units_to_create
