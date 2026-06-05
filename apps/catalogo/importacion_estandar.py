"""Importador del estándar de equipamiento INACAP desde planillas Excel."""

import re
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from apps.catalogo.models import Asignatura, TipoEquipo


@dataclass
class ResumenImportacion:
    """Resumen serializable de una importación de estándar de equipamiento."""

    tipos_creados: int = 0
    tipos_actualizados: int = 0
    asignaturas_creadas: int = 0
    vinculos_creados: int = 0
    advertencias: list[str] = field(default_factory=list)
    _tipos_actualizados_ids: set[int] = field(default_factory=set, repr=False)

    def marcar_tipo_actualizado(self, tipo_equipo: TipoEquipo) -> None:
        if tipo_equipo.pk not in self._tipos_actualizados_ids:
            self.tipos_actualizados += 1
            self._tipos_actualizados_ids.add(tipo_equipo.pk)

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data.pop("_tipos_actualizados_ids", None)
        return data


EQUIPOS_COLUMNAS = {
    "nombre": 0,
    "especificacion": 1,
    "cantidad": 2,
    "codigo_ps": 5,
    "observacion": 6,
}
RELACION_COLUMNAS = {
    "codigo": 0,
    "semestre": 1,
    "asignatura": 2,
    "nombre": 3,
    "especificacion": 4,
    "cantidad": 5,
}


def limpiar_texto(value: Any) -> str:
    """Convierte una celda a texto tolerando None, espacios y saltos de línea."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalizar_clave_nombre(value: Any) -> str:
    return limpiar_texto(value).casefold()


def parsear_cantidad(value: Any, resumen: ResumenImportacion, contexto: str) -> int:
    """Parsea cantidades positivas; una celda vacía vale 0 y no advierte."""
    if value is None or limpiar_texto(value) == "":
        return 0
    if isinstance(value, int):
        return max(value, 0)
    if isinstance(value, float):
        return max(int(value), 0)

    text = limpiar_texto(value).replace(",", ".")
    try:
        return max(int(Decimal(text)), 0)
    except (InvalidOperation, ValueError):
        resumen.advertencias.append(
            f"{contexto}: cantidad no numérica '{limpiar_texto(value)}'; se usa 0."
        )
        return 0


def _celda(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _buscar_tipo_por_clave(nombre: str, tipo_seguimiento: str) -> TipoEquipo | None:
    nombre_normalizado = normalizar_clave_nombre(nombre)
    candidatos = TipoEquipo.objects.filter(tipo_seguimiento=tipo_seguimiento)
    for candidato in candidatos:
        if normalizar_clave_nombre(candidato.nombre) == nombre_normalizado:
            return candidato
    return None


def _observaciones(codigo_ps: str, observacion: str) -> str:
    partes = []
    if codigo_ps:
        partes.append(f"cód PS: {codigo_ps}")
    if observacion:
        partes.append(observacion)
    return "\n".join(partes)


def _fusionar_observaciones(actual: str, nueva: str) -> str:
    if not nueva:
        return actual
    if not actual:
        return nueva
    lineas_actuales = {linea.strip() for linea in actual.splitlines() if linea.strip()}
    nuevas_lineas = [
        linea.strip()
        for linea in nueva.splitlines()
        if linea.strip() and linea.strip() not in lineas_actuales
    ]
    if not nuevas_lineas:
        return actual
    return "\n".join([actual, *nuevas_lineas])


def _registrar_tipo(
    *,
    nombre: str,
    especificacion: str,
    cantidad: int,
    tipo_seguimiento: str,
    observaciones: str,
    resumen: ResumenImportacion,
) -> TipoEquipo:
    tipo_equipo = _buscar_tipo_por_clave(nombre, tipo_seguimiento)
    if tipo_equipo is None:
        resumen.tipos_creados += 1
        return TipoEquipo.objects.create(
            nombre=limpiar_texto(nombre),
            especificacion=especificacion,
            cantidad_necesaria=cantidad,
            tipo_seguimiento=tipo_seguimiento,
            observaciones=observaciones,
        )

    campos_actualizados: list[str] = []
    if not tipo_equipo.especificacion and especificacion:
        tipo_equipo.especificacion = especificacion
        campos_actualizados.append("especificacion")
    if cantidad > tipo_equipo.cantidad_necesaria:
        tipo_equipo.cantidad_necesaria = cantidad
        campos_actualizados.append("cantidad_necesaria")

    observaciones_fusionadas = _fusionar_observaciones(
        tipo_equipo.observaciones, observaciones
    )
    if observaciones_fusionadas != tipo_equipo.observaciones:
        tipo_equipo.observaciones = observaciones_fusionadas
        campos_actualizados.append("observaciones")

    if campos_actualizados:
        tipo_equipo.save(update_fields=campos_actualizados)
        resumen.marcar_tipo_actualizado(tipo_equipo)
    return tipo_equipo


def _procesar_hoja_tipos(
    workbook,
    hoja: str,
    tipo_seguimiento: str,
    resumen: ResumenImportacion,
) -> dict[str, list[TipoEquipo]]:
    indice: dict[str, list[TipoEquipo]] = {}
    if hoja not in workbook.sheetnames:
        resumen.advertencias.append(f"No se encontró la hoja '{hoja}'.")
        return indice

    worksheet = workbook[hoja]
    for numero_fila, row in enumerate(
        worksheet.iter_rows(min_row=5, values_only=True), start=5
    ):
        nombre = limpiar_texto(_celda(row, EQUIPOS_COLUMNAS["nombre"]))
        if not nombre:
            if any(limpiar_texto(value) for value in row):
                resumen.advertencias.append(
                    f"{hoja} fila {numero_fila}: fila omitida por falta de nombre."
                )
            continue

        cantidad = parsear_cantidad(
            _celda(row, EQUIPOS_COLUMNAS["cantidad"]),
            resumen,
            f"{hoja} fila {numero_fila}",
        )
        tipo_equipo = _registrar_tipo(
            nombre=nombre,
            especificacion=limpiar_texto(
                _celda(row, EQUIPOS_COLUMNAS["especificacion"])
            ),
            cantidad=cantidad,
            tipo_seguimiento=tipo_seguimiento,
            observaciones=_observaciones(
                limpiar_texto(_celda(row, EQUIPOS_COLUMNAS["codigo_ps"])),
                limpiar_texto(_celda(row, EQUIPOS_COLUMNAS["observacion"])),
            ),
            resumen=resumen,
        )
        indice.setdefault(normalizar_clave_nombre(nombre), []).append(tipo_equipo)
    return indice


def _resolver_tipos_relacion(
    nombre: str,
    especificacion: str,
    indice_tipos: dict[str, list[TipoEquipo]],
) -> list[TipoEquipo]:
    candidatos = indice_tipos.get(normalizar_clave_nombre(nombre), [])
    if len(candidatos) <= 1 or not especificacion:
        return candidatos

    especificacion_normalizada = limpiar_texto(especificacion).casefold()
    coincidencias = [
        tipo
        for tipo in candidatos
        if limpiar_texto(tipo.especificacion).casefold() == especificacion_normalizada
    ]
    return coincidencias or candidatos


def _procesar_relaciones(
    workbook,
    resumen: ResumenImportacion,
    indice_tipos: dict[str, list[TipoEquipo]],
) -> None:
    hoja = "Relación Asig"
    if hoja not in workbook.sheetnames:
        resumen.advertencias.append(f"No se encontró la hoja '{hoja}'.")
        return

    worksheet = workbook[hoja]
    for numero_fila, row in enumerate(
        worksheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        codigo = limpiar_texto(_celda(row, RELACION_COLUMNAS["codigo"])).upper()
        nombre_asignatura = limpiar_texto(_celda(row, RELACION_COLUMNAS["asignatura"]))
        nombre_equipo = limpiar_texto(_celda(row, RELACION_COLUMNAS["nombre"]))
        if not any((codigo, nombre_asignatura, nombre_equipo)):
            continue
        if not codigo:
            resumen.advertencias.append(
                f"{hoja} fila {numero_fila}: relación omitida por falta "
                "de código de asignatura."
            )
            continue
        if not nombre_asignatura:
            resumen.advertencias.append(
                f"{hoja} fila {numero_fila}: relación omitida por falta "
                "de nombre de asignatura."
            )
            continue
        if not nombre_equipo:
            resumen.advertencias.append(
                f"{hoja} fila {numero_fila}: relación omitida por falta "
                "de nombre de equipo."
            )
            continue

        asignatura, creada = Asignatura.objects.get_or_create(
            codigo=codigo, defaults={"nombre": nombre_asignatura}
        )
        if creada:
            resumen.asignaturas_creadas += 1
        elif not asignatura.nombre and nombre_asignatura:
            asignatura.nombre = nombre_asignatura
            asignatura.save(update_fields=["nombre"])

        tipos = _resolver_tipos_relacion(
            nombre_equipo,
            limpiar_texto(_celda(row, RELACION_COLUMNAS["especificacion"])),
            indice_tipos,
        )
        if not tipos:
            resumen.advertencias.append(
                f"{hoja} fila {numero_fila}: no se encontró TipoEquipo "
                f"para '{nombre_equipo}'."
            )
            continue

        for tipo_equipo in set(tipos):
            if not tipo_equipo.asignaturas.filter(pk=asignatura.pk).exists():
                tipo_equipo.asignaturas.add(asignatura)
                resumen.vinculos_creados += 1


def _combinar_indices(
    *indices: dict[str, list[TipoEquipo]],
) -> dict[str, list[TipoEquipo]]:
    combinado: dict[str, list[TipoEquipo]] = {}
    for indice in indices:
        for clave, tipos in indice.items():
            combinado.setdefault(clave, []).extend(tipos)
    return combinado


@transaction.atomic
def importar_estandar(file_o_path, *, planes=None) -> ResumenImportacion:
    """Importa un estándar INACAP, sin crear unidades físicas ni stock real.

    La deduplicación de TipoEquipo usa nombre normalizado + tipo de seguimiento.
    Si un tipo ya existe, solo se completa la especificación vacía y se conserva
    la mayor cantidad_necesaria encontrada entre importaciones.
    """
    del planes  # Reservado para agrupar futuras cargas por plan de estudio.

    workbook = load_workbook(file_o_path, data_only=True, read_only=True)
    resumen = ResumenImportacion()
    indice_equipos = _procesar_hoja_tipos(
        workbook,
        "Equipos",
        TipoEquipo.TipoSeguimiento.SERIE,
        resumen,
    )
    indice_insumos = _procesar_hoja_tipos(
        workbook,
        "Insumos",
        TipoEquipo.TipoSeguimiento.GRANEL,
        resumen,
    )
    _procesar_relaciones(
        workbook,
        resumen,
        _combinar_indices(indice_equipos, indice_insumos),
    )
    workbook.close()
    return resumen
