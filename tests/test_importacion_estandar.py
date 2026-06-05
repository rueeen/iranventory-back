from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from rest_framework.test import APIClient

from apps.catalogo.importacion_estandar import importar_estandar
from apps.catalogo.models import Asignatura, TipoEquipo
from apps.cuentas.models import Usuario
from apps.inventario.models import Unidad


@pytest.fixture
def panolero(db, django_user_model):
    return django_user_model.objects.create_user(
        username="panolero-estandar",
        password="clave-123",
        rol=Usuario.Rol.PANOLERO,
    )


@pytest.fixture
def alumno(db, django_user_model):
    return django_user_model.objects.create_user(
        username="alumno-estandar",
        password="clave-123",
        rol=Usuario.Rol.ALUMNO,
    )


@pytest.fixture
def api_client():
    return APIClient()


def construir_xlsx_estandar() -> BytesIO:
    workbook = Workbook()
    resumen = workbook.active
    resumen.title = "Resumen"
    equipos = workbook.create_sheet("Equipos")
    insumos = workbook.create_sheet("Insumos")
    workbook.create_sheet("Equipos Año")
    relacion = workbook.create_sheet("Relación Asig")

    for sheet in (equipos, insumos):
        sheet.append(["Plan de estudio"])
        sheet.append(["Título"])
        sheet.append([])
        sheet.append(
            [
                "equipo",
                "especificación técnica",
                "cantidad",
                "valor unitario - uf (c/iva)",
                "valor total - uf (c/iva)",
                "cód ps",
                "Observación",
                "valor total proyectado",
            ]
        )

    equipos.append(
        [
            " Osciloscopio\nDigital ",
            "100 MHz\n2 canales",
            2,
            1,
            2,
            "000000000000003463",
            "Referencia docente",
        ]
    )
    equipos.append(["Multímetro", "True RMS", "no aplica", None, None, "PS-2"])
    equipos.append([None, "Sin nombre", 3])

    insumos.append(["Cable banana", "Rojo/negro", 10, None, None, "PS-3"])
    insumos.append(["Protoboard", "830 puntos", 4, None, None, "PS-4"])
    insumos.append(["  ", "Sin nombre", 1])

    relacion.append(["Plan de estudio"])
    relacion.append([])
    relacion.append(
        [
            "código",
            "semestre",
            "nombre asignatura",
            "nombre",
            "especificación técnica",
            "cantidad",
        ]
    )
    relacion.append(
        ["O401", "I", "Electrónica I", "Osciloscopio Digital", "100 MHz 2 canales", 1]
    )
    relacion.append(["O401", "I", "Electrónica I", "Cable banana", "Rojo/negro", 6])
    relacion.append(["O402", "II", "Mediciones", "Multímetro", "True RMS", 1])
    relacion.append(["O403", "II", "Sin equipo", None, "", 1])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


@pytest.mark.django_db
def test_importar_estandar_crea_tipos_asignaturas_vinculos_y_advertencias():
    resumen = importar_estandar(construir_xlsx_estandar())

    osciloscopio = TipoEquipo.objects.get(nombre="Osciloscopio Digital")
    cable = TipoEquipo.objects.get(nombre="Cable banana")

    assert resumen.tipos_creados == 4
    assert resumen.asignaturas_creadas == 2
    assert resumen.vinculos_creados == 3
    assert osciloscopio.tipo_seguimiento == TipoEquipo.TipoSeguimiento.SERIE
    assert cable.tipo_seguimiento == TipoEquipo.TipoSeguimiento.GRANEL
    assert osciloscopio.cantidad_necesaria == 2
    assert cable.cantidad_necesaria == 10
    assert osciloscopio.stock_granel == 0
    assert cable.stock_granel == 0
    assert Unidad.objects.count() == 0
    assert "cód PS: 000000000000003463" in osciloscopio.observaciones
    assert Asignatura.objects.get(codigo="O401") in osciloscopio.asignaturas.all()
    assert Asignatura.objects.get(codigo="O401") in cable.asignaturas.all()
    assert any("falta de nombre" in advertencia for advertencia in resumen.advertencias)
    assert any(
        "cantidad no numérica" in advertencia for advertencia in resumen.advertencias
    )


@pytest.mark.django_db
def test_importar_estandar_es_idempotente_y_conserva_mayor_cantidad():
    importar_estandar(construir_xlsx_estandar())
    resumen = importar_estandar(construir_xlsx_estandar())

    assert TipoEquipo.objects.count() == 4
    assert Asignatura.objects.count() == 2
    assert resumen.tipos_creados == 0
    assert resumen.vinculos_creados == 0
    assert TipoEquipo.objects.get(nombre="Osciloscopio Digital").cantidad_necesaria == 2

    workbook_file = construir_xlsx_estandar()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen"
    equipos = workbook.create_sheet("Equipos")
    insumos = workbook.create_sheet("Insumos")
    workbook.create_sheet("Equipos Año")
    relacion = workbook.create_sheet("Relación Asig")
    for target in (equipos, insumos):
        target.append([])
        target.append([])
        target.append([])
        target.append(["equipo", "especificación técnica", "cantidad"])
    equipos.append(["Osciloscopio Digital", "", 1])
    relacion.append([])
    relacion.append([])
    relacion.append(["código", "semestre", "nombre asignatura", "nombre"])
    workbook_file.close()
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    importar_estandar(output)

    assert TipoEquipo.objects.get(nombre="Osciloscopio Digital").cantidad_necesaria == 2


@pytest.mark.django_db
def test_endpoint_importar_estandar_devuelve_resumen(api_client, panolero):
    api_client.force_authenticate(user=panolero)
    archivo = SimpleUploadedFile(
        "estandar.xlsx",
        construir_xlsx_estandar().getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response = api_client.post(
        "/api/catalogo/importar-estandar/", {"archivo": archivo}, format="multipart"
    )

    assert response.status_code == 200
    assert response.data["tipos_creados"] == 4
    assert response.data["asignaturas_creadas"] == 2
    assert TipoEquipo.objects.count() == 4
    assert Unidad.objects.count() == 0


@pytest.mark.django_db
def test_endpoint_importar_estandar_rechaza_no_xlsx(api_client, panolero):
    api_client.force_authenticate(user=panolero)
    archivo = SimpleUploadedFile(
        "estandar.txt", b"contenido", content_type="text/plain"
    )

    response = api_client.post(
        "/api/catalogo/importar-estandar/", {"archivo": archivo}, format="multipart"
    )

    assert response.status_code == 400
    assert "archivo" in response.data


@pytest.mark.django_db
def test_endpoint_importar_estandar_rechaza_alumno(api_client, alumno):
    api_client.force_authenticate(user=alumno)
    archivo = SimpleUploadedFile(
        "estandar.xlsx",
        construir_xlsx_estandar().getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response = api_client.post(
        "/api/catalogo/importar-estandar/", {"archivo": archivo}, format="multipart"
    )

    assert response.status_code == 403
